"""test_output.py - Test cases for the output module."""

import zipfile
import pytest
import pandas as pd
import openpyxl
from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock, patch
from datetime import datetime, date, timedelta

from log_keeper.output import (
    launches_to_excel,
    backup_launches_collection,
    launches_to_db,
    update_launches_collection,
    update_aircraft_info,
    fill_log_sheet,
)

TEMPLATE_FIXTURE = Path(__file__).parent / "fixtures" / "2965D_260214_ZE633.xlsx"
# The repo's docs/ template is the empty scaffold (blank header cells).
EMPTY_TEMPLATE = (Path(__file__).parent.parent / "docs"
                  / "2965D_YYMMDD_ZEXXX.xltx")


@pytest.fixture
def sample_launches_df():
    """Create a sample launches DataFrame."""
    return pd.DataFrame({
        'Date': [pd.Timestamp('2025-01-23'), pd.Timestamp('2025-01-23')],
        'Aircraft': ['ZE123', 'ZE456'],
        'AircraftCommander': ['John Doe', 'Jane Smith'],
        '2ndPilot': ['Bob Wilson', 'Alice Brown'],
        'Duty': ['Training', 'Check'],
        'TakeOffTime': [pd.Timestamp('2025-01-23 10:00:00'), pd.Timestamp('2025-01-23 11:00:00')],
        'LandingTime': [pd.Timestamp('2025-01-23 10:30:00'), pd.Timestamp('2025-01-23 11:45:00')],
        'FlightTime': [30, 45],
        'SPC': [1, 2],
        'PLF': [True, False],
        'P1': [True, False],
        'P2': [False, True]
    })


@pytest.fixture
def sample_aircraft_info():
    """Create a sample aircraft info DataFrame."""
    return pd.DataFrame({
        'Date': [pd.Timestamp('2025-01-23')],
        'Aircraft': ['ZE123'],
        'Launches After': [100],
        'Hours After': [120]
    })


@pytest.fixture
def mock_db():
    """Create a mock Database object."""
    mock = MagicMock()
    mock.db = MagicMock()
    mock.launches_collection = "launches"
    mock.aircraft_info_collection = "aircraft_info"
    return mock


def test_launches_to_excel(tmp_path, sample_launches_df):
    """Test saving launches to Excel."""
    output_path = tmp_path / "MASTER_LOG.xlsx"

    # Test successful save
    with patch('xlsxwriter.Workbook') as mock_workbook:
        mock_book = MagicMock()
        mock_workbook.return_value = mock_book
        mock_book.add_format.return_value = MagicMock()
        launches_to_excel(sample_launches_df, output_path)

    # Test fallback to temp file
    with patch('xlsxwriter.Workbook') as mock_workbook:
        mock_workbook.side_effect = [Exception("Test error"), MagicMock()]
        launches_to_excel(sample_launches_df, output_path)


def test_backup_launches_collection(mock_db):
    """Test backing up launches collection."""
    # Test when backup doesn't exist
    mock_db.db.list_collection_names.return_value = []
    backup_launches_collection(mock_db)
    mock_db.get_launches_collection.assert_called_once()

    # Test when backup exists
    mock_db.db.list_collection_names.return_value = [
        f"launches_{datetime.today().strftime('%y%m%d')}"
    ]
    backup_launches_collection(mock_db)
    mock_db.db.drop_collection.assert_called_once()


def test_launches_to_db(mock_db, sample_launches_df):
    """Test saving launches to database."""
    launches_to_db(sample_launches_df, mock_db)
    mock_db.get_launches_collection.assert_called()
    mock_db.get_launches_collection.return_value.insert_many.assert_called_once()


def test_update_launches_collection(mock_db, sample_launches_df):
    """Test updating launches collection."""
    # Test with non-empty DataFrame
    collection = mock_db.get_launches_collection.return_value
    collection.count_documents.return_value = 2
    update_launches_collection(sample_launches_df, mock_db)
    collection.bulk_write.assert_called_once()
    collection.insert_many.assert_called_once()

    # Test with empty DataFrame
    update_launches_collection(pd.DataFrame(), mock_db)
    assert collection.bulk_write.call_count == 1  # Should not increase


def test_update_aircraft_info(mock_db, sample_aircraft_info):
    """Test updating aircraft info."""
    # Test with non-empty DataFrame
    collection = mock_db.get_aircraft_info_collection.return_value
    collection.count_documents.return_value = 1
    update_aircraft_info(sample_aircraft_info, mock_db)
    collection.bulk_write.assert_called_once()
    collection.insert_many.assert_called_once()

    # Test with empty DataFrame
    update_aircraft_info(pd.DataFrame(), mock_db)
    assert collection.bulk_write.call_count == 1  # Should not increase


def test_fill_log_sheet():
    """fill_log_sheet pre-fills the header cells and preserves the workbook."""
    src = TEMPLATE_FIXTURE.read_bytes()
    out = fill_log_sheet(src, "ZE683", launches_bf=1234,
                         hours_bf_minutes=74070, sheet_date=date(2026, 6, 30))

    # Every other zip part (form controls, drawings, tables) is preserved.
    n_src = len(zipfile.ZipFile(BytesIO(src)).infolist())
    n_out = len(zipfile.ZipFile(BytesIO(out)).infolist())
    assert n_out == n_src

    # The download is a real workbook, not a template Excel opens as a copy.
    content_types = zipfile.ZipFile(BytesIO(out)).read(
        "[Content_Types].xml").decode("utf-8")
    assert "sheet.main+xml" in content_types
    assert "template.main+xml" not in content_types

    # The four header cells are filled (C4 as an exact [h]:mm duration).
    ws = openpyxl.load_workbook(BytesIO(out))["2965D"]
    assert ws["F2"].value == "ZE683"
    assert ws["O2"].value == datetime(2026, 6, 30)
    assert ws["C4"].value == timedelta(minutes=74070)
    assert ws["C5"].value == 1234

    # The hidden sheets the ingest pipeline relies on still parse.
    with pd.ExcelFile(BytesIO(out)) as xls:
        assert {"FORMATTED", "_AIRCRAFT"} <= set(xls.sheet_names)


def test_fill_log_sheet_blank_brought_forward():
    """With no brought-forward data, C4/C5 are left blank."""
    out = fill_log_sheet(EMPTY_TEMPLATE.read_bytes(), "ZE557",
                         launches_bf=None, hours_bf_minutes=None,
                         sheet_date=date(2026, 6, 30))
    ws = openpyxl.load_workbook(BytesIO(out))["2965D"]
    assert ws["F2"].value == "ZE557"
    assert ws["C4"].value is None
    assert ws["C5"].value is None
