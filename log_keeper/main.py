"""log_keeper.py
661 VGS - Collate all log sheets (2965D) into one master log."""

# Get packages.
from pathlib import Path
import pandas as pd
import warnings
from datetime import datetime
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers

def ingest_log_sheet(file_path):
    """
    Extract data from an excel log sheet.
    Output a pandas dataframe.
    """
    # Constants.
    SHEET_NAME = "FORMATTED"
    
    # Read from the log sheet.
    raw_df = pd.read_excel(
    file_path, 
    sheet_name=SHEET_NAME,
    dtype={
        'AircraftCommander': 'string',
        '2ndPilot': 'string',
        'Duty': 'string',
        'TakeOffTime': 'datetime64[ns]',
        'LandingTime': 'datetime64[ns]',
        'FlightTime': 'UInt16',
        'SPC': 'UInt8',
        'PLF': 'bool',
        'Aircraft': 'string',
        'Date': 'string'
        }
    )
    
    return raw_df


def sanitise_log_sheets(log_sheet_df):
    """Filter and replace data in the master log dataframe."""

    # Filter the log sheets to remove AircraftCommander "0"
    log_sheet_df = log_sheet_df[log_sheet_df.AircraftCommander != "0"]

    # Change GIC to GIF.
    log_sheet_df.loc[log_sheet_df['Duty'] == 'GIC', 'Duty'] = 'GIF'

    # Add SCT into QGI and U/T duties e.g. "QGI" -> "SCT QGI"
    log_sheet_df.loc[log_sheet_df['Duty'] == 'U/T', 'Duty'] = 'SCT U/T'
    log_sheet_df.loc[log_sheet_df['Duty'] == 'QGI', 'Duty'] = 'SCT QGI'
    
    # Change aircraft commander and second pilot to Upper Case.
    log_sheet_df.loc[:, 'AircraftCommander'] = log_sheet_df['AircraftCommander'].str.title()
    log_sheet_df.loc[:, '2ndPilot'] = log_sheet_df['2ndPilot'].str.title()

    # Change date to DD/MM/YYYY format.
    log_sheet_df.loc[:, 'Date'] = log_sheet_df['Date'].str[:10]

    # Sort by takeofftime.
    log_sheet_df = log_sheet_df.sort_values(by="TakeOffTime", ascending=True, na_position="first")

    # Rename 2ndPilot to SecondPilot.
    log_sheet_df.rename(columns={"2ndPilot": "SecondPilot"}, inplace=True)

    # Change FlightTime to be a timedelta instead of integer.
    log_sheet_df['FlightTime'] = pd.to_timedelta(log_sheet_df['FlightTime'], unit='m').dt.floor('min')

    return log_sheet_df


def collate_log_sheets(dir_path):
    """Collate all log sheets into a single dataframe using the path to the log sheet directory."""
    # Get the directory contents.
    FILE_NAME = "2965D_*.xlsx"
    dir_contents = dir_path.glob(f"{FILE_NAME}")
    log_sheet_files = [x for x in dir_contents if x.is_file()]

    # Extract data from each log sheet.
    # Constants.
    N_COLS = 10

    for i_file, file_path in enumerate(log_sheet_files):
        # Get the log sheet data.
        try:
            this_sheet_df = ingest_log_sheet(file_path)
            if i_file == 0:
                # Create a new dataframe based on the first file ingest.
                log_sheet_df = this_sheet_df
            else:
                # Append to the master dataframe.
                log_sheet_df = pd.concat([log_sheet_df, this_sheet_df], ignore_index=True)
        except Exception as e:
            if file_path.name != "2965D_YYMMDD_ZEXXX.xlsx":
                warnings.warn(f"Log sheet invalid: {file_path.name}", RuntimeWarning)
                print(e)


    collated_df = sanitise_log_sheets(log_sheet_df)

    return collated_df


def master_log_to_excel(master_log, output_file_path):
    """Save the master log dataframe to an excel table."""

    # Sheet the table should be inserted into.
    SHEET_NAME = "Master Log"

    # Create the excel file and table.
    with pd.ExcelWriter(output_file_path, engine='openpyxl', 
                            mode='a', if_sheet_exists='replace') as writer:
        master_log.to_excel(
            writer,
            sheet_name=SHEET_NAME,
            index=False,
            header=True,
        )

        # Access the workbook and worksheet.
        workbook = writer.book
        worksheet = workbook[SHEET_NAME]

        # Get the range of the data.
        min_row = worksheet.min_row
        max_row = worksheet.max_row
        min_column = worksheet.min_column
        max_column = worksheet.max_column
        data_range = f"A1:{get_column_letter(max_column)}{max_row}"

        # Create the table.
        table = Table(displayName="MasterLog", ref=data_range)

        # Apply the table style.
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)

        # Set column widths.
        dim_holder = DimensionHolder(worksheet=worksheet)

        # Get the length of the column title and set the column width accordingly
        for col in range(min_column, max_column + 1):
            column_letter = get_column_letter(col)
            cell = worksheet[column_letter + "1"]  # Assuming the column title is in the first row
            title_length = len(cell.value)  # Get the length of the title.
            column_width = title_length + 6  # Adjust if needed for better fit
            dim_holder[column_letter] = ColumnDimension(worksheet, min=col, max=col, width=column_width)

        worksheet.column_dimensions = dim_holder

        # Set the number format for 'TakeOffTime' and 'LandingTime' columns.
        time_format = 'h:mm'

        for column_name in ['TakeOffTime', 'LandingTime']:
            column_num = master_log.columns.get_loc(column_name) + 1
            for col in worksheet.iter_cols(min_row=min_row + 1, max_row=max_row,
                                            min_col=column_num, max_col=column_num):
                # For each cell in the column.
                for cell in col:
                    cell.number_format = time_format

        # Set date format for 'FlightTime' column.
        duration_format = '[h]:mm'
        column_num = master_log.columns.get_loc('FlightTime') + 1
        for col in worksheet.iter_cols(min_row=min_row + 1, max_row=max_row,
                                            min_col=column_num, max_col=column_num):
            # For each cell in the column.
            for cell in col:
                cell.number_format = duration_format

        # Set date format for 'Date' column.
        date_format = numbers.FORMAT_DATE_DDMMYY
        column_num = master_log.columns.get_loc('Date') + 1
        for col in worksheet.iter_cols(min_row=min_row + 1, max_row=max_row,
                                            min_col=column_num, max_col=column_num):
            # For each cell in the column.
            for cell in col:
                cell.number_format = date_format

        # Add the table to the worksheet.
        worksheet.add_table(table)


def main():
    # Initial comment.
    print("viking-log-keeper: Starting...")

    # Get the file path.
    root_dir = Path.home()

    # Path to the sharepoint directory.
    SHAREPOINT_DIR = Path(
        root_dir, 
        "Royal Air Force Air Cadets",
        "661 VGS - RAF Kirknewton - 661 Documents",
    )

    # If the path does not exist, its probably the squadron laptop.
    # There is definitely a better way of doing this.
    if SHAREPOINT_DIR.exists() == False:
        SHAREPOINT_DIR = Path(
            root_dir, 
            "Onedrive - Royal Air Force Air Cadets",
            "661 Documents",
        )

    # Path to the log sheets directory.
    LOG_SHEETS_DIR = Path(
        SHAREPOINT_DIR, 
        "Log Sheets"
    )

    # Output file path.
    OUTPUT_FILE = Path(
        SHAREPOINT_DIR,
        "Stats",
        "Dashboard.xlsx"
    )

    # Create a dataframe of all log sheets.
    master_log = collate_log_sheets(LOG_SHEETS_DIR)

    # Save the master log to excel.
    master_log_to_excel(master_log, OUTPUT_FILE)

    # Print success message.
    print("viking-log-keeper: Success!")


if __name__ == "__main__":
    main()